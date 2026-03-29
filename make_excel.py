from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ---- Sheet 1: 周五上午班 ----
ws1 = wb.active
ws1.title = "周五上午班-儿童画"

headers = ["序号", "姓名", "电话号码", "备注"]
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(color="FFFFFF", bold=True)

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

data_am = [
    (1, "周瞻", "18726218681", ""),
    (2, "王雨清", "18714807613", ""),
    (3, "孙嘉祺", "17355019076", ""),
    (4, "郑辉辉", "18755096419", ""),
    (5, "刘世义", "13721005979", ""),
    (6, "徐青松", "668400", "短号"),
    (7, "赵恒毅", "13956286330", ""),
    (8, "王悦", "13149575906", ""),
    (9, "宋潇", "13705509737", ""),
    (10, "房子然", "15212086027", ""),
    (11, "鄂鸣超", "13093315899", ""),
]

for row_idx, row_data in enumerate(data_am, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws1.cell(row=row_idx, column=col_idx, value=val)

# ---- Sheet 2: 周五下午班 ----
ws2 = wb.create_sheet("周五下午班-儿童画")

for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

data_pm = [
    (1, "丁淑贤", "13955094554", ""),
    (2, "王昕楠", "15055039191", ""),
    (3, "徐青松", "668400", "短号"),
    (4, "胡碧清", "1822663430***", "号码不完整"),
    (5, "孙甜甜", "13695501958", ""),
    (6, "关蕾璐", "13955044769", ""),
    (7, "王嘉欣", "15212081212", ""),
    (8, "宣改维", "13737217127", ""),
    (9, "黄***", "15855087832", "姓名不完整"),
    (10, "鄂鸣超", "13093315899", ""),
    (11, "赵恒毅", "13956286330", ""),
    (12, "刘振晨", "13955042623", ""),
    (13, "柏子涵", "18714805980", ""),
    (14, "陈辉平", "15358062366", ""),
    (15, "王欣燕", "18355009790", ""),
    (16, "陈浩然", "1872667071***", "号码不完整"),
]

for row_idx, row_data in enumerate(data_pm, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=val)

# ---- Sheet 3: 5月19日晚上班 ----
ws3 = wb.create_sheet("5月19日晚上班")

for col, h in enumerate(headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

data_night = [
    (1, "吴敏怡", "18098378574", ""),
    (2, "罗忠博", "15856658186", ""),
    (3, "王思琪", "15385095358", ""),
    (4, "郭幸玲", "13155017674", ""),
    (5, "申咏鑫", "13083319702", ""),
    (6, "谭凯辰", "13365704961", ""),
    (7, "吉光辰", "13355087855", ""),
    (8, "徐思乔", "15255229451", ""),
    (9, "李乐心", "", "无号码"),
    (10, "范康意", "18110751168", ""),
    (11, "陆爱迪", "", "无号码"),
    (12, "萧紫兮", "15212005111 / 15212009300", "两个号码"),
    (13, "陈潇阳", "13866944191", ""),
    (14, "蔡经阁", "18155044187", ""),
]

for row_idx, row_data in enumerate(data_night, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=val)

# ---- Sheet 4: 5月19日下午班 ----
ws4 = wb.create_sheet("5月19日下午班")

for col, h in enumerate(headers, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

data_pm2 = [
    (1, "翁云峰", "18112409158", ""),
    (2, "王悦", "13516416322", ""),
    (3, "陈教畔", "18755098614", ""),
    (4, "陈正鸿", "13855075444", ""),
    (5, "徐兵器", "13173217771", "原号15855085853已改"),
    (6, "柏德标", "18455084345", ""),
    (7, "顾成缘", "13355502832", ""),
    (8, "汪皓珩", "13965988761", ""),
    (9, "王宝玲", "15956082226", ""),
    (10, "陈晓莉", "", "无号码"),
    (11, "周璐", "18726218681", ""),
    (12, "王婷洁", "13955099171", ""),
    (13, "翁正峰", "18712409158", ""),
    (14, "陈晓渔", "18755098614", ""),
    (15, "刘珊", "13470705827", ""),
    (16, "杜斯法", "13865825677", ""),
]

for row_idx, row_data in enumerate(data_pm2, 2):
    for col_idx, val in enumerate(row_data, 1):
        ws4.cell(row=row_idx, column=col_idx, value=val)

# 设置列宽
for ws in [ws1, ws2, ws3, ws4]:
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20

output_path = "/home/elliot/.openclaw/workspace/儿童画学员名单.xlsx"
wb.save(output_path)
print(f"已保存到: {output_path}")
